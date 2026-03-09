import io
import os
import logging
from fastapi import HTTPException, UploadFile
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

async def convert_pdf_to_excel(file: UploadFile):
    """
    Convert a PDF file to an Excel (.xlsx) workbook.
    - Extracts all tables from each page using pdfplumber.
    - Each page with tables gets its own sheet.
    - Pages with no tables fall back to plain text rows.
    """
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are accepted. Please upload a .pdf file.",
        )

    logger.info(f"Converting to Excel: {file.filename}")

    pdf_bytes = await file.read()
    if len(pdf_bytes) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        wb = openpyxl.Workbook()
        # Remove the default empty sheet
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            total_pages = len(pdf.pages)
            sheets_created = 0

            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()

                if tables:
                    for t_idx, table in enumerate(tables):
                        sheet_name = (
                            f"P{page_num}" if len(tables) == 1
                            else f"P{page_num}_T{t_idx + 1}"
                        )
                        ws = wb.create_sheet(title=sheet_name)
                        sheets_created += 1

                        for row_idx, row in enumerate(table, start=1):
                            for col_idx, cell_value in enumerate(row, start=1):
                                cell = ws.cell(
                                    row=row_idx,
                                    column=col_idx,
                                    value=str(cell_value).strip() if cell_value is not None else "",
                                )
                                # Style the first row as a header
                                if row_idx == 1:
                                    cell.fill = header_fill
                                    cell.font = header_font
                                    cell.alignment = Alignment(horizontal="center")

                        # Auto-fit column widths
                        for col in ws.columns:
                            max_len = max(
                                (len(str(c.value)) if c.value else 0 for c in col), default=0
                            )
                            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                                max_len + 4, 60
                            )
                else:
                    # No tables — fall back to plain text rows
                    text = page.extract_text()
                    if text and text.strip():
                        ws = wb.create_sheet(title=f"Page {page_num}")
                        sheets_created += 1
                        for line in text.splitlines():
                            if line.strip():
                                ws.append([line.strip()])

        # If absolutely nothing was extracted, add an info sheet
        if sheets_created == 0:
            ws = wb.create_sheet(title="No Data")
            ws.append(["No extractable tables or text were found in this PDF."])
            ws.append(["This may be a scanned/image-based PDF. Try OCR first."])

        # Serialize workbook to bytes
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_bytes = xlsx_buffer.getvalue()

    except Exception as e:
        logger.error(f"Excel conversion failed: {e}")
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

    base_name = os.path.splitext(file.filename)[0]
    output_filename = f"{base_name}.xlsx"
    
    return xlsx_bytes, output_filename
